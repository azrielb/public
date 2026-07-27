import sys
import os
import csv
import openpyxl


def xlsx_to_csv(input_path, output_path=None, sheet_name=None):
    if output_path is None:
        base = os.path.splitext(input_path)[0]
        output_path = base + ".csv"

    wb = openpyxl.load_workbook(input_path, data_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            print(f"Error: sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
            sys.exit(1)
        ws = wb[sheet_name]
    else:
        ws = wb.active

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])

    print(f"Saved: {output_path}  ({ws.max_row} rows, {ws.max_column} cols)")


def print_usage():
    print("Usage: xls2csv.py <input.xlsx> [output.csv] [--sheet <name>]")
    print("  If output is omitted, uses the same name with .csv extension.")
    print("  If --sheet is omitted, converts the active (first) sheet.")


if __name__ == "__main__":
    args = sys.argv[1:]
    if not args or args[0] in ("-h", "--help"):
        print_usage()
        sys.exit(0)

    input_file = args[0]
    output_file = None
    sheet = None

    i = 1
    while i < len(args):
        if args[i] == "--sheet" and i + 1 < len(args):
            sheet = args[i + 1]
            i += 2
        elif not args[i].startswith("--"):
            output_file = args[i]
            i += 1
        else:
            print(f"Unknown option: {args[i]}")
            print_usage()
            sys.exit(1)

    if not os.path.isfile(input_file):
        print(f"Error: file not found: {input_file}")
        sys.exit(1)

    xlsx_to_csv(input_file, output_file, sheet)
